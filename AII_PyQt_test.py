import os
import sys
import pandas as pd
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon, QFont, QColor, QIntValidator, QPixmap
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel,
                             QFileDialog, QVBoxLayout, QHBoxLayout, QWidget,
                             QMessageBox, QTableWidget, QTableWidgetItem,
                             QRadioButton, QGroupBox, QSplitter, QTabWidget,
                             QSizePolicy, QComboBox, QLineEdit, QScrollArea,
                             QGridLayout, QTreeWidget, QTreeWidgetItem)

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.font_manager as fm
import AII, FCE_AII_new, util
import chardet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


diff_map_3 = {'特别困难': 3, '一般困难': 2, '不困难': 1}
diff_map_4 = {'特别困难': 4, '困难': 3, '一般困难': 2, '不困难': 1}


class MatplotlibCanvas(FigureCanvas):

    def __init__(self, parent=None, width=5, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super(MatplotlibCanvas, self).__init__(self.fig)
        self.setParent(parent)

    def plot_poverty_stats(self, df):
        self.axes.clear()
        if '算法认定结果' in df.columns:
            stats = df['算法认定结果'].value_counts()
            font = fm.FontProperties(fname="C:/Windows/Fonts/msyh.ttc",
                                     size=12)
            bars = self.axes.bar(stats.index,
                                 stats.values,
                                 color='#3f72af',
                                 edgecolor='black')
            for bar in bars:
                height = bar.get_height()
                self.axes.text(bar.get_x() + bar.get_width() / 2,
                               height + 10,
                               f'{int(height)}',
                               ha='center',
                               va='bottom',
                               fontproperties=font)
            self.axes.set_title("各困难等级人数统计", fontproperties=font)
            self.axes.set_ylabel("人数", fontproperties=font)
            self.axes.set_xlabel("困难等级", fontproperties=font)
            self.axes.grid(axis='y', linestyle='--', alpha=0.4)
            self.axes.set_facecolor('#fefefe')
        self.draw()


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("家庭经济困难学生精准认定工具")
        self.setGeometry(100, 100, 1280, 860)
        self.dataset_file = ''
        self.zbtx_file = ''
        self.tmp_dir = './tmp'
        self.selected_algorithm = "AII"
        self.df_std = None
        # self.canvas = MatplotlibCanvas(self)
        self.is_simple_view = True  # 添加视图状态标志
        self.current_page = 1  # 当前页码
        self.rows_per_page = 10  # 每页显示行数
        self.total_pages = 0  # 总页数
        self.initUI()

    def initUI(self):
        # 设置全局字体为微软雅黑
        font = QFont("Microsoft YaHei", 9)
        QApplication.instance().setFont(font)

        self.setStyleSheet("""
            QWidget { background-color: #eaf0f6; font-family: 'Microsoft YaHei'; font-size: 10pt; }
            QLabel { color: #1f2d3d; }
            QPushButton {
                background-color: #3f72af; color: white; border: none;
                padding: 6px 12px; border-radius: 4px;
            }
            QPushButton:hover { background-color: #2b5d8c; }
            QGroupBox {
                border: 1px solid #ccc; border-radius: 6px;
                margin-top: 10px; font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 10px; padding: 0 3px;
            }
            QTableWidget {
                background-color: white; border: 1px solid #ccc;
                gridline-color: #e0e0e0;
            }
            QHeaderView::section {
                background-color: #f3f6f9; padding: 6px; font-weight: bold;
                border: 1px solid #ccc;
            }
            QLineEdit {
                padding: 5px; border: 1px solid #ccc; border-radius: 4px; background-color: white;
            }
            QTreeWidget {
                background-color: white; border: 1px solid #ccc;
                border-radius: 4px; padding: 5px;
            }
            QTreeWidget::item {
                padding: 5px;
            }
            QTreeWidget::item:hover {
                background-color: #e8f4f8;
            }
            QTreeWidget::item:selected {
                background-color: #3f72af;
                color: white;
            }
        """)

        file_group = QGroupBox("文件选择")
        file_layout = QVBoxLayout(file_group)
        file_layout.setContentsMargins(10, 20, 10, 10)
        file_layout.setSpacing(5)

        # 待认定数据文件
        h_dataset = QHBoxLayout()
        h_dataset.addWidget(QLabel("待认定数据文件:"))
        h_dataset.addStretch()
        btn_dataset = QPushButton("浏览...")
        btn_dataset.clicked.connect(self.select_dataset_file)
        h_dataset.addWidget(btn_dataset)
        file_layout.addLayout(h_dataset)

        self.label_dataset_name = QLabel("未选择文件")
        self.label_dataset_name.setStyleSheet(
            "font-size: 8pt; margin-bottom: 5px;")
        file_layout.addWidget(self.label_dataset_name)

        # 指标体系文件
        h_zbtx = QHBoxLayout()
        h_zbtx.addWidget(QLabel("指标体系文件:"))
        h_zbtx.addStretch()
        btn_zbtx = QPushButton("浏览...")
        btn_zbtx.clicked.connect(self.select_zbtx_file)
        h_zbtx.addWidget(btn_zbtx)
        file_layout.addLayout(h_zbtx)

        self.label_zbtx_name = QLabel("未选择文件")
        self.label_zbtx_name.setStyleSheet(
            "font-size: 8pt; margin-bottom: 5px;")
        file_layout.addWidget(self.label_zbtx_name)

        self.radio_aii = QRadioButton("AII")
        self.radio_aii.setChecked(True)
        self.radio_fce_aii = QRadioButton("FCE-AII")
        self.radio_aii.toggled.connect(self.toggle_algorithm)
        self.radio_fce_aii.toggled.connect(self.toggle_algorithm)

        # 新增：算法执行方式
        self.radio_mode_recognition = QRadioButton("困难度认定")
        self.radio_mode_recognition.setChecked(True)
        self.radio_mode_comparison = QRadioButton("人工-算法认定结果比对")
        self.radio_mode_comparison.toggled.connect(
            self.check_mode_requirements)

        # 新增：困难度认定比重分配
        self.spin_special = QLineEdit("10")
        self.spin_difficult = QLineEdit("20")
        self.spin_general = QLineEdit("30")
        self.spin_not_difficult = QLineEdit("40")

        # 设置验证器，只允许输入0-100的数字
        validator = QIntValidator(0, 100)
        self.spin_special.setValidator(validator)
        self.spin_difficult.setValidator(validator)
        self.spin_general.setValidator(validator)
        self.spin_not_difficult.setValidator(validator)

        self.label_ratio_source = QLabel("来源: 手动输入")
        self.label_ratio_source.setStyleSheet(
            "font-size: 8pt; margin-bottom: 5px;")

        btn_run = QPushButton("运行算法")
        btn_save = QPushButton("保存结果")
        btn_run.clicked.connect(self.run_algorithm)
        btn_save.clicked.connect(self.save_result)

        algo_box = QGroupBox("🧠 选择算法")
        algo_layout = QVBoxLayout()
        algo_layout.setContentsMargins(10, 20, 10, 10)
        algo_layout.addWidget(self.radio_aii)
        algo_layout.addWidget(self.radio_fce_aii)
        algo_box.setLayout(algo_layout)

        # 新增：算法执行方式 GroupBox
        mode_box = QGroupBox("⚙️ 算法执行方式")
        mode_layout = QVBoxLayout()
        mode_layout.setContentsMargins(10, 20, 10, 10)
        mode_layout.addWidget(self.radio_mode_recognition)
        mode_layout.addWidget(self.radio_mode_comparison)
        mode_box.setLayout(mode_layout)

        # 新增：困难度认定比重分配 GroupBox
        ratio_box = QGroupBox("📊 困难度认定比重分配 (%)")
        ratio_layout = QGridLayout()
        ratio_layout.setContentsMargins(10, 20, 10, 10)
        ratio_layout.addWidget(QLabel("特别困难:"), 0, 0)
        ratio_layout.addWidget(self.spin_special, 0, 1)
        ratio_layout.addWidget(QLabel("困难:"), 1, 0)
        ratio_layout.addWidget(self.spin_difficult, 1, 1)
        ratio_layout.addWidget(QLabel("一般困难:"), 2, 0)
        ratio_layout.addWidget(self.spin_general, 2, 1)
        ratio_layout.addWidget(QLabel("不困难:"), 3, 0)
        ratio_layout.addWidget(self.spin_not_difficult, 3, 1)
        ratio_layout.addWidget(self.label_ratio_source, 4, 0, 1, 2)
        ratio_box.setLayout(ratio_layout)

        control_layout = QVBoxLayout()
        control_layout.setSpacing(12)
        control_layout.addWidget(file_group)  # 使用新的文件选择组
        # control_layout.addWidget(self.label_output)
        # control_layout.addWidget(btn_select_output)
        control_layout.addWidget(algo_box)
        control_layout.addWidget(mode_box)  # 添加模式选择
        control_layout.addWidget(ratio_box)  # 添加比重分配
        control_layout.addWidget(btn_run)
        control_layout.addWidget(btn_save)
        control_layout.addStretch(1)
        self.add_fupin_label(control_layout)
        control_widget = QWidget()
        control_widget.setLayout(control_layout)
        control_widget.setSizePolicy(QSizePolicy.Preferred,
                                     QSizePolicy.Expanding)

        self.table_input = QTableWidget()
        # self.table_output = QTableWidget()

        # 创建指标体系树
        self.tree_indicator = QTreeWidget()
        self.tree_indicator.setHeaderLabel("指标体系结构")
        self.tree_indicator.setMinimumWidth(250)
        self.tree_indicator.setStyleSheet(
            "QTreeWidget::item { margin: 0px 0px;}")  # 行距调整

        # 在输入表格上方添加切换按钮
        self.btn_toggle_view = QPushButton("切换到详细视图")
        self.btn_toggle_view.setStyleSheet("background-color: #3f72af;")
        self.btn_toggle_view.clicked.connect(self.toggle_view)

        # 创建分页控件
        self.page_label = QLabel("第1页")
        self.total_page_label = QLabel("共0页")

        self.prev_btn = QPushButton("上一页")
        self.next_btn = QPushButton("下一页")
        self.prev_btn.setEnabled(False)
        self.next_btn.setEnabled(False)
        self.prev_btn.clicked.connect(self.prev_page)
        self.next_btn.clicked.connect(self.next_page)

        self.rows_combo = QComboBox()
        self.rows_combo.addItems(["10", "20", "50", "100"])
        self.rows_combo.setCurrentText("10")
        self.rows_combo.currentTextChanged.connect(self.change_rows_per_page)

        # 在输入表格上方添加切换按钮
        self.btn_toggle_view = QPushButton("切换到详细视图")
        self.btn_toggle_view.setStyleSheet("background-color: #3f72af;")
        self.btn_toggle_view.clicked.connect(self.toggle_view)

        # 添加自定义页码输入框和跳转按钮
        self.page_input = QLineEdit()
        self.page_input.setFixedWidth(50)
        self.page_input.setPlaceholderText("页码")
        self.page_input.setValidator(QIntValidator(1, 9999))  # 只允许输入正整数

        self.jump_btn = QPushButton("跳转")
        self.jump_btn.clicked.connect(self.jump_to_page)

        # 分页控件布局
        pagination_layout = QHBoxLayout()
        pagination_layout.addWidget(self.btn_toggle_view)  # 切换视图按钮
        pagination_layout.addWidget(QLabel("每页行数:"))
        pagination_layout.addWidget(self.rows_combo)
        pagination_layout.addStretch()
        pagination_layout.addWidget(self.prev_btn)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(QLabel("/"))
        pagination_layout.addWidget(self.total_page_label)
        pagination_layout.addWidget(self.next_btn)
        pagination_layout.addWidget(QLabel("跳转到:"))
        pagination_layout.addWidget(self.page_input)  # 添加页码输入框
        pagination_layout.addWidget(self.jump_btn)  # 添加跳转按钮

        # 将分页控件添加到表格容器
        table_container = QWidget()
        table_layout = QVBoxLayout()
        table_layout.addLayout(pagination_layout)  # 添加分页控件
        table_layout.addWidget(self.table_input)
        table_container.setLayout(table_layout)

        # 创建水平分割器,左侧显示指标体系树,右侧显示数据表格
        data_splitter = QSplitter(Qt.Horizontal)
        data_splitter.addWidget(self.tree_indicator)
        data_splitter.addWidget(table_container)
        data_splitter.setStretchFactor(0, 1)  # 树形控件占1份
        data_splitter.setStretchFactor(1, 3)  # 表格占3份

        self.tab_widget = QTabWidget()
        self.tab_widget.addTab(data_splitter, "📥 困难认定数据展示")  # 修改为使用分割器

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(control_widget)
        splitter.addWidget(self.tab_widget)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 2)

        layout = QVBoxLayout()
        layout.addWidget(splitter)
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def toggle_view(self):
        """切换简洁视图和详细视图"""
        self.is_simple_view = not self.is_simple_view

        if self.is_simple_view:
            self.btn_toggle_view.setText("切换到详细视图")
            self.btn_toggle_view.setStyleSheet("background-color: #3f72af;")
        else:
            self.btn_toggle_view.setText("切换到简洁视图")
            self.btn_toggle_view.setStyleSheet("background-color: #3f72af;")

        # 如果数据已加载，重新显示表格
        if hasattr(self, 'df_full') and self.df_full is not None:
            self.display_table_data(self.table_input, self.df_full)

    def display_table_data(self, table_widget, df):
        """根据当前视图模式显示数据"""
        # 应用视图过滤
        if self.is_simple_view:
            # 简洁视图只保留三列
            simple_columns = [
                "学校所在地", "学校", "学院", "学生姓名", "学号", "实际认定结果", "算法认定结果"
            ]

            # 确保这些列存在
            available_columns = [
                col for col in simple_columns if col in df.columns
            ]

            df = df[available_columns]

        # 清理未命名列
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        # TODO !!!!!!!注意：正常为4 3 2 1，目前一些学校没有4个等级，调整为3个等级处理
        # df['实际认定结果'].value_counts() # 检查实际认定结果的等级数量
        cnt = df['实际认定结果'].nunique()
        assert cnt in (3, 4), "实际认定结果等级数量异常！"
        diff_map = diff_map_4 if cnt == 4 else diff_map_3
        # 添加结果比较列
        if '实际认定结果' in df.columns and '算法认定结果' in df.columns:
            df['结果是否一致'] = df.apply(
                lambda row: '一致' if row['实际认定结果'] == row['算法认定结果'] else '不一致',
                axis=1)
            # 差异性列内容：差一级、差两级以上 二选一
            df['差异性'] = df.apply(lambda row: '无差异'
                                 if row['实际认定结果'] == row['算法认定结果'] else
                                 ('差两级以上' if abs(
                                     diff_map.get(row['实际认定结果'], 0) - diff_map.
                                     get(row['算法认定结果'], 0)) >= 2 else '差一级'),
                                 axis=1)

        # 计算分页
        total_rows = df.shape[0]
        # 修复除零错误
        if total_rows == 0:
            self.total_pages = 1
        else:
            self.total_pages = max(1, (total_rows + self.rows_per_page - 1) //
                                   self.rows_per_page)
        self.total_page_label.setText(f"共{self.total_pages}页")

        # 修正当前页码
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages

        # 计算当前页的数据范围
        start_row = (self.current_page - 1) * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, total_rows)
        page_df = df.iloc[start_row:end_row]

        # 显示表格数据
        table_widget.setRowCount(page_df.shape[0])
        table_widget.setColumnCount(page_df.shape[1])
        table_widget.setHorizontalHeaderLabels(page_df.columns)

        font = QFont("Microsoft YaHei", 10, QFont.Normal)
        table_widget.setFont(font)

        # 检查是否同时存在 "实际认定结果" 和 "算法认定结果" 两列
        has_actual_result = "实际认定结果" in page_df.columns
        has_algorithm_result = "算法认定结果" in page_df.columns

        for row in range(page_df.shape[0]):
            # 默认行背景颜色
            row_color = QColor(255, 255, 255)  # 白色背景

            # 如果同时存在 "实际认定结果" 和 "算法认定结果"，对比两列的值
            if has_actual_result and has_algorithm_result:
                actual_result = str(page_df.iloc[row]["实际认定结果"])
                algorithm_result = str(page_df.iloc[row]["算法认定结果"])

                # 根据一致性设置整行背景颜色
                if actual_result == algorithm_result:
                    row_color = QColor(220, 255, 220)  # 绿色背景
                else:
                    row_color = QColor(255, 220, 220)  # 粉色背景

            for col in range(page_df.shape[1]):
                # 获取单元格内容
                content = str(page_df.iat[row, col])
                is_numeric = False
                try:
                    # 尝试转换为数值
                    num = float(content)
                    is_numeric = True
                    # 检查是否为整数
                    if num.is_integer():
                        # 如果是整数，去掉小数点
                        content = str(int(num))
                    else:
                        # 如果是浮点数，保留四位小数
                        content = f"{num:.4f}"
                except ValueError:
                    # 不是数值，保持原样
                    pass

                item = QTableWidgetItem(content)
                item.setFont(font)

                # 设置整行背景颜色
                item.setBackground(row_color)

                # 设置对齐方式
                if is_numeric:
                    item.setTextAlignment(Qt.AlignRight
                                          | Qt.AlignVCenter)  # 数值右对齐
                else:
                    item.setTextAlignment(Qt.AlignLeft
                                          | Qt.AlignVCenter)  # 非数值左对齐

                table_widget.setItem(row, col, item)

        table_widget.resizeColumnsToContents()
        table_widget.resizeRowsToContents()
        table_widget.horizontalHeader().setMinimumHeight(30)
        table_widget.verticalHeader().setDefaultSectionSize(32)
        self.update_pagination()

    def jump_to_page(self):
        """跳转到指定页码"""
        if not self.page_input.text().isdigit():
            QMessageBox.warning(self, "无效输入", "请输入有效的页码！")
            return

        target_page = int(self.page_input.text())

        if target_page < 1 or target_page > self.total_pages:
            QMessageBox.warning(self, "无效页码",
                                f"超出页码范围 (1-{self.total_pages})! 请输入有效页码。")
            return

        self.current_page = target_page
        if hasattr(self, 'df_full') and self.df_full is not None:
            self.display_table_data(self.table_input, self.df_full)

    def update_pagination(self):
        """更新分页控件状态"""
        self.page_label.setText(f"第{self.current_page}页")
        self.total_page_label.setText(f"共{self.total_pages}页")
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < self.total_pages)

    def change_rows_per_page(self, text):
        """改变每页显示行数"""
        self.rows_per_page = int(text)
        self.current_page = 1
        if hasattr(self, 'df_full') and self.df_full is not None:
            self.display_table_data(self.table_input, self.df_full)

    def prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            if hasattr(self, 'df_full') and self.df_full is not None:
                self.display_table_data(self.table_input, self.df_full)

    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            if hasattr(self, 'df_full') and self.df_full is not None:
                self.display_table_data(self.table_input, self.df_full)

    def toggle_algorithm(self):
        if self.radio_aii.isChecked():
            self.selected_algorithm = "AII"
        elif self.radio_fce_aii.isChecked():
            self.selected_algorithm = "FCEAII"

    def check_mode_requirements(self):
        """检查模式要求"""
        if self.radio_mode_comparison.isChecked():
            if self.dataset_file:
                try:
                    df = pd.read_csv(self.dataset_file)
                    if '实际认定结果' not in df.columns:
                        QMessageBox.critical(
                            self, "错误",
                            "当前选择'人工-算法认定结果比对'模式，但数据文件中未包含'实际认定结果'列！\n请切换模式或重新选择文件。"
                        )
                        self.radio_mode_recognition.setChecked(True)
                except Exception as e:
                    QMessageBox.warning(self, "警告", f"读取文件失败: {str(e)}")

    def select_dataset_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择数据文件", '',
                                              "CSV Files (*.csv)")
        if file:
            self.dataset_file = file
            self.label_dataset_name.setText(os.path.basename(file))
            self.label_dataset_name.setToolTip(file)

            try:
                df = pd.read_csv(file)
                self.df_full = df  # 保存完整数据
                self.display_table_data(self.table_input, df)  # 使用新方法显示

                # 检查是否包含实际认定结果并自动填充比重
                if '实际认定结果' in df.columns:
                    total_count = len(df)
                    if total_count > 0:
                        counts = df['实际认定结果'].value_counts()

                        # 计算百分比并四舍五入
                        p_special = int(
                            round((counts.get('特别困难', 0) / total_count) * 100))
                        p_difficult = int(
                            round((counts.get('困难', 0) / total_count) * 100))
                        p_general = int(
                            round((counts.get('一般困难', 0) / total_count) * 100))
                        p_not_difficult = int(
                            round((counts.get('不困难', 0) / total_count) * 100))

                        # 调整总和为100
                        current_sum = p_special + p_difficult + p_general + p_not_difficult
                        diff = 100 - current_sum

                        if diff != 0:
                            # 将差值加到数量最多的类别上，以减少相对误差
                            raw_counts = {
                                'special': counts.get('特别困难', 0),
                                'difficult': counts.get('困难', 0),
                                'general': counts.get('一般困难', 0),
                                'not_difficult': counts.get('不困难', 0)
                            }
                            max_key = max(raw_counts, key=raw_counts.get)

                            if max_key == 'special':
                                p_special += diff
                            elif max_key == 'difficult':
                                p_difficult += diff
                            elif max_key == 'general':
                                p_general += diff
                            else:
                                p_not_difficult += diff

                        # 更新UI
                        self.spin_special.setText(f"{p_special}")
                        self.spin_difficult.setText(f"{p_difficult}")
                        self.spin_general.setText(f"{p_general}")
                        self.spin_not_difficult.setText(f"{p_not_difficult}")
                        self.label_ratio_source.setText("来源: 自动获取 (基于实际认定结果)")
                        self.label_ratio_source.setStyleSheet(
                            "color: green; font-weight: bold; font-size: 12px;"
                        )
                else:
                    # 如果没有实际认定结果，且处于比对模式，则报错
                    if self.radio_mode_comparison.isChecked():
                        QMessageBox.critical(
                            self, "错误",
                            "当前选择'人工-算法认定结果比对'模式，但数据文件中未包含'实际认定结果'列！")
                        self.radio_mode_recognition.setChecked(True)

                    self.label_ratio_source.setText("来源: 手动输入")
                    self.label_ratio_source.setStyleSheet(
                        "color: gray; font-size: 12px;")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件失败: {str(e)}")

    def select_zbtx_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择指标体系文件", '',
                                              "CSV Files (*.csv)")
        if file:
            self.zbtx_file = file
            self.label_zbtx_name.setText(os.path.basename(file))
            self.label_zbtx_name.setToolTip(file)

            # 加载并显示指标体系树
            self.load_indicator_tree(file)

    def load_indicator_tree(self, csv_file):

        """加载指标体系CSV文件并以树状结构显示"""
        try:
            with open(csv_file, 'rb') as f:
                raw_data = f.read()
                detected_encoding = chardet.detect(raw_data)['encoding']

            df = pd.read_csv(csv_file, encoding=detected_encoding)
            self.tree_indicator.clear()

            # 使用字典来存储已创建的节点,避免重复
            level1_dict = {}  # indicator_1 节点
            level2_dict = {}  # indicator_2 节点

            # 首先收集所有实际分数以确定颜色范围
            all_scores = df['score'].dropna().tolist()
            if all_scores:
                min_score = min(all_scores)
                max_score = max(all_scores)
                score_range = max_score - min_score if max_score > min_score else 1
            else:
                min_score = 0
                max_score = 1
                score_range = 1

            for _, row in df.iterrows():
                ind1 = str(row['indicator_1'])
                ind2 = str(row['indicator_2'])
                ind3 = str(row['indicator_3'])
                score = row.get('score', '')
                normalized_score = row.get('normalized_score', '')

                # 创建或获取一级节点
                if ind1 not in level1_dict:
                    level1_item = QTreeWidgetItem(self.tree_indicator)
                    level1_item.setText(0, ind1)
                    level1_item.setExpanded(True)  # 默认展开
                    # 设置一级节点字体加粗
                    font1 = level1_item.font(0)
                    font1.setBold(True)
                    level1_item.setFont(0, font1)
                    level1_dict[ind1] = level1_item
                else:
                    level1_item = level1_dict[ind1]

                # 创建或获取二级节点
                key2 = f"{ind1}_{ind2}"
                if key2 not in level2_dict:
                    level2_item = QTreeWidgetItem(level1_item)
                    level2_item.setText(0, ind2)
                    level2_item.setExpanded(True)  # 默认展开
                    level2_dict[key2] = level2_item
                else:
                    level2_item = level2_dict[key2]

                # 创建三级节点(叶子节点)
                level3_item = QTreeWidgetItem(level2_item)
                # 显示三级指标名称和分数信息
                if pd.notna(score):
                    if pd.notna(normalized_score):
                        level3_item.setText(
                            0,
                            f"{ind3} [分数: {score}, 归一化: {normalized_score:.2f}]"
                        )
                    else:
                        level3_item.setText(0, f"{ind3} [分数: {score}]")

                    # 根据实际分数设置蓝色渐变背景色
                    # 计算当前分数在整体范围中的比例
                    score_val = float(score)
                    norm_ratio = (score_val - min_score) / score_range

                    # 使用蓝色渐变: 从浅蓝(低分)到深蓝(高分)
                    # 浅蓝色 (250, 250, 255) 到 深蓝色 (13, 71, 161)
                    r = int(250 - (250 - 13) * norm_ratio)
                    g = int(250 - (250 - 71) * norm_ratio)
                    b = int(255 - (255 - 161) * norm_ratio)

                    bg_color = QColor(r, g, b)
                    level3_item.setBackground(0, bg_color)

                    # 当背景较深时，文字使用白色以提高可读性
                    if norm_ratio > 0.5:
                        level3_item.setForeground(0, QColor(255, 255, 255))
                else:
                    level3_item.setText(0, ind3)

        except Exception as e:
            QMessageBox.warning(self, "警告", f"加载指标体系失败: {str(e)}")

    def create_image_widget(self, img_path, font, base_size):
        """创建单个图片显示组件"""
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(10, 10, 10, 10)
        container_layout.setSpacing(10)

        # 创建图片显示区域
        image_label = QLabel()
        image_label.setAlignment(Qt.AlignCenter)

        # 保存原始图片以供缩放
        pixmap = QPixmap(img_path)
        image_label.original_pixmap = pixmap

        # 设置初始大小
        image_label.setMinimumSize(base_size, base_size)

        # 添加标题标签
        caption = QLabel(os.path.splitext(os.path.basename(img_path))[0])
        caption.setFont(font)
        caption.setAlignment(Qt.AlignCenter)
        caption.setWordWrap(True)

        # 将组件添加到容器中
        container_layout.addWidget(image_label)
        container_layout.addWidget(caption)

        # 保存 label 引用以便后续更新
        container.image_label = image_label
        return container

    def update_image_sizes(self, scroll_area, containers, cols):
        """更新所有图片的大小"""
        # 获取可用空间
        available_width = scroll_area.viewport().width()
        padding = 40  # 总边距
        spacing = 20  # 图片间距

        # 计算单个图片的目标大小
        image_width = (available_width - (padding * 2) - (spacing *
                                                          (cols - 1))) // cols
        image_width = max(200, min(600, image_width))  # 限制最小/最大尺寸

        # 更新每个图片容器
        for container in containers:
            if hasattr(container, 'image_label'):
                label = container.image_label
                if hasattr(label, 'original_pixmap'):
                    # 保持纵横比缩放图片
                    scaled_pixmap = label.original_pixmap.scaled(
                        image_width, image_width, Qt.KeepAspectRatio,
                        Qt.SmoothTransformation)
                    label.setPixmap(scaled_pixmap)
                    # 更新最小尺寸以保持布局稳定
                    label.setMinimumSize(image_width, image_width)

    class ImageScrollArea(QScrollArea):

        def __init__(self, parent=None):
            super().__init__(parent)
            self.resize_timer = QTimer(self)
            self.resize_timer.setSingleShot(True)
            self.resize_timer.setInterval(100)
            self.containers = []
            self.cols = 1

        def set_resize_callback(self, callback):
            self.resize_timer.timeout.connect(callback)

        def resizeEvent(self, event):
            super().resizeEvent(event)
            self.resize_timer.start()

    def plot_multiple_images(self, image_dir, pattern_prefix):
        # 创建自定义滚动区域
        scroll = self.ImageScrollArea()
        scroll.setWidgetResizable(True)

        image_container = QWidget()
        grid_layout = QGridLayout()
        grid_layout.setSpacing(20)

        # 获取所有匹配的图片文件
        files = sorted([
            f for f in os.listdir(image_dir)
            if f.startswith(pattern_prefix) and f.endswith('.png')
        ])

        # 根据图片数量确定列数
        num_images = len(files)
        if num_images <= 2:
            cols = num_images
        elif num_images <= 4:
            cols = 2
        else:
            cols = 3

        scroll.cols = cols

        # 设置字体和基础大小
        font = QFont("Microsoft YaHei", 10)
        base_size = 400

        # 创建所有图片容器
        containers = []
        row, col = 0, 0
        for fname in files:
            img_path = os.path.join(image_dir, fname)
            container = self.create_image_widget(img_path, font, base_size)
            containers.append(container)

            grid_layout.addWidget(container, row, col)
            col += 1
            if col >= cols:
                col = 0
                row += 1

        # 设置布局
        grid_layout.setRowStretch(row, 1)
        grid_layout.setColumnStretch(cols, 1)
        image_container.setLayout(grid_layout)
        scroll.setWidget(image_container)

        # 保存容器引用并设置回调
        scroll.containers = containers
        scroll.set_resize_callback(
            lambda: self.update_image_sizes(scroll, containers, cols))

        # 初始更新图片大小
        self.update_image_sizes(scroll, containers, cols)

        return scroll

    def add_image_tab(self, tab_widget):
        """
        ✅ 新增方法：将图像结果作为标签页添加到界面中。
        应在 run_AII() 和 run_FCE_AII() 中调用：
            self.add_image_tab(tab_widget)
        并确保 tab_widget 是 QTabWidget 实例
        """
        if self.tmp_dir and os.path.isdir(self.tmp_dir):
            image_tab = self.plot_multiple_images(
                self.tmp_dir, self.selected_algorithm + "_")
            tab_widget.addTab(image_tab, "📈 认定结果总览")

    def run_algorithm(self):
        # 从输入框获取路径
        # self.dataset_file 和 self.zbtx_file 已经在选择文件时设置

        if not all([self.dataset_file, self.zbtx_file, self.tmp_dir]):
            QMessageBox.warning(self, "⚠️ 警告", "请完整选择文件和输出目录！")
            return

        # 检查比重分配总和是否为100%
        try:
            p_special = int(self.spin_special.text())
            p_difficult = int(self.spin_difficult.text())
            p_general = int(self.spin_general.text())
            p_not_difficult = int(self.spin_not_difficult.text())

            total = p_special + p_difficult + p_general + p_not_difficult
            if total != 100:
                QMessageBox.warning(self, "警告",
                                    f"困难度认定比重总和必须为100%！当前总和: {total}%")
                return
        except ValueError:
            QMessageBox.warning(self, "警告", "请输入有效的比重数值！")
            return

        try:
            if self.selected_algorithm == "AII":
                self.run_AII()
            elif self.selected_algorithm == "FCEAII":
                self.run_FCE_AII()
        except Exception as e:
            QMessageBox.critical(self, "❌ 错误", str(e))

    def add_fupin_label(self, layout):
        self.fupin_label = QLabel("🎯 精准资助")
        self.fupin_label.setAlignment(Qt.AlignCenter)
        self.fupin_label.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        self.fupin_label.setStyleSheet(
            "color: #2b5d8c; margin-top: 30px; margin-bottom: 10px;")
        layout.addWidget(self.fupin_label)

    def run_AII(self):
        # df_std = RRDA.run_RRDA(self.dataset_file)
        with open(self.dataset_file, 'rb') as f:
            raw_data = f.read()
            detected_encoding = chardet.detect(raw_data)['encoding']
        df_std = pd.read_csv(self.dataset_file, encoding=detected_encoding)
        with open(self.zbtx_file, 'rb') as f:
            raw_data = f.read()
            detected_encoding = chardet.detect(raw_data)['encoding']
        df_idx = pd.read_csv(self.zbtx_file, encoding=detected_encoding)

        # 获取用户输入的比重
        poverty_perc = {
            '特别困难': int(self.spin_special.text()) / 100.0,
            '困难': int(self.spin_difficult.text()) / 100.0,
            '一般困难': int(self.spin_general.text()) / 100.0,
            '不困难': int(self.spin_not_difficult.text()) / 100.0
        }

        # 如果是比对模式，确保有实际认定结果列
        if self.radio_mode_comparison.isChecked():
            if '实际认定结果' not in df_std.columns:
                QMessageBox.critical(
                    self, "错误", "当前选择'人工-算法认定结果比对'模式，但数据文件中未包含'实际认定结果'列！")
                return

        A_I, A_II = util.get_AI_AII(df_idx)
        # RRDA_AI_indicater_1=RRDA_AI_indicater_1)
        # RRDA_score = A_II[RRDA_AI_indicater_1]['家庭经济困难度指数'] * A_I[
        #     RRDA_AI_indicater_1] * 100
        self.df_std = AII.AII(df_std,
                              save=False,
                              df_idx=df_idx,
                              work_dir=self.tmp_dir,
                              poverty_perc=poverty_perc,
                              patch=True)
        self.df_full = self.df_std  # 保存完整数据
        self.display_table_data(self.table_input, self.df_std)  # 使用新方法显示

        # 仅在比对模式下显示一致性
        if self.radio_mode_comparison.isChecked():
            self.display_acc()

        QMessageBox.information(self, "成功", "AII算法完成")
        self.add_image_tab(self.tab_widget)

    def run_FCE_AII(self):
        # df_std = RRDA.run_RRDA(self.dataset_file)
        with open(self.dataset_file, 'rb') as f:
            raw_data = f.read()
            detected_encoding = chardet.detect(raw_data)['encoding']
        df_std = pd.read_csv(self.dataset_file, encoding=detected_encoding)
        self.df_std = FCE_AII_new.FCE_AII(df_std,
                                          self.zbtx_file,
                                          save=False,
                                          work_dir=self.tmp_dir,
                                          with_RRDA=False)
        self.df_full = self.df_std  # 保存完整数据
        self.display_table_data(self.table_input, self.df_std)  # 使用新方法显示
        self.display_acc()
        QMessageBox.information(self, "成功", "FCE-AII算法完成")
        self.add_image_tab(self.tab_widget)

    def display_acc(self):
        if self.df_std is not None and '实际认定结果' in self.df_std.columns:
            acc = util.calc_acc(self.df_std, '算法认定结果', '实际认定结果')
            if acc is not None:
                QMessageBox.information(self, "一致性计算",
                                        f"一致性为 {acc * 100:.2f}%")

    # def save_results(self):
    #     if self.df_std is not None:
    #         # 创建DataFrame的副本以避免修改原始数据
    #         df_to_save = self.df_std.copy()

    #         # 如果存在实际认定结果和算法认定结果，添加比较列
    #         if '实际认定结果' in df_to_save.columns and '算法认定结果' in df_to_save.columns:
    #             # 添加结果比较列
    #             df_to_save['结果是否一致'] = df_to_save.apply(
    #                 lambda row: '一致' if row['实际认定结果'] == row['算法认定结果'] else '不一致',
    #                 axis=1
    #             )

    #         # 添加差异性列
    #         df_to_save['差异性'] = df_to_save.apply(
    #             lambda row: '差两级以上' if (
    #                 (row['实际认定结果'] == '不困难' and row['算法认定结果'] == '特别困难') or
    #                 (row['实际认定结果'] == '特别困难' and row['算法认定结果'] == '不困难')
    #             ) else '差一级',
    #             axis=1
    #         )

    #         # 设置默认文件名为 "困难认定结果.csv"
    #         default_file_name = os.path.join(os.path.expanduser("~"),
    #                                          "困难认定结果.csv")

    #         # 弹出文件管理器选择保存路径
    #         output_file, _ = QFileDialog.getSaveFileName(
    #             self,
    #             "选择保存文件路径",
    #             default_file_name,  # 默认文件名
    #             "CSV Files (*.csv)"  # 文件类型过滤
    #         )

    #         if output_file:  # 如果用户选择了路径
    #             # 确保文件名以 .csv 结尾
    #             if not output_file.endswith(".csv"):
    #                 output_file += ".csv"

    #             # 保存文件
    #             df_to_save.to_csv(output_file,
    #                                encoding='utf_8_sig',
    #                                index=False)
    #             QMessageBox.information(self, "保存成功", f"结果已保存至: {output_file}")
    #         else:
    #             QMessageBox.information(self, "取消保存", "未选择保存路径")
    #     else:
    #         QMessageBox.warning(self, "警告", "请先运行算法")

    def save_result(self):
        if self.df_std is None:
            QMessageBox.warning(self, "警告", "请先运行算法")
            return

        # 1. 准备要保存的数据 (直接复用 display_table_data 的逻辑)
        # 注意：这里我们使用完整数据 self.df_full (即 self.df_std)，而不是视图过滤后的数据
        # 如果用户希望保存"看到的每页数据"，通常也是指向保存完整结果。
        # 为了保证"一致性"和"差异性"列存在，我们需要重新应用相同的逻辑。

        df_to_save = self.df_std.copy()

        # 清理未命名列
        df_to_save = df_to_save.loc[:, ~df_to_save.columns.str.
                                    contains('^Unnamed')]

        # 复用 display_table_data 中的计算逻辑
        if '实际认定结果' in df_to_save.columns and '算法认定结果' in df_to_save.columns:
            cnt = df_to_save['实际认定结果'].nunique()
            assert cnt in (3, 4), "实际认定结果等级数量异常！"
            diff_map = diff_map_4 if cnt == 4 else diff_map_3
            df_to_save['结果是否一致'] = df_to_save.apply(
                lambda row: '一致' if row['实际认定结果'] == row['算法认定结果'] else '不一致',
                axis=1)
            df_to_save['差异性'] = df_to_save.apply(
                lambda row: '无差异'
                if row['实际认定结果'] == row['算法认定结果'] else ('差两级以上' if abs(
                    diff_map.get(row['实际认定结果'], 0) - diff_map.get(
                        row['算法认定结果'], 0)) >= 2 else '差一级'),
                axis=1)

        # 2. 获取保存路径
        default_file_name = os.path.join(os.path.expanduser("~"),
                                         "困难认定结果.xlsx")
        output_file, _ = QFileDialog.getSaveFileName(self, "选择保存文件路径",
                                                     default_file_name,
                                                     "Excel Files (*.xlsx)")

        if not output_file:
            QMessageBox.information(self, "取消保存", "未选择保存路径")
            return

        if not output_file.endswith(".xlsx"):
            output_file += ".xlsx"

        # 3. 保存并应用颜色样式 (逻辑与 display_table_data 的颜色逻辑对应)
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_to_save.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']

                # 定义颜色 (与 UI 中的颜色对应: 绿色=一致, 粉色/红色=不一致)
                # display_table_data: QColor(220, 255, 220) -> hex DCFFDC
                # display_table_data: QColor(255, 220, 220) -> hex FFDCDC
                green_fill = PatternFill(start_color='DCFFDC',
                                         end_color='DCFFDC',
                                         fill_type='solid')
                red_fill = PatternFill(start_color='FFDCDC',
                                       end_color='FFDCDC',
                                       fill_type='solid')
                # 差异大的用更深的红色，方便区分 (可选，display_table_data 中只用了两种背景色，这里为了更细致保留了深红)
                dark_red_fill = PatternFill(start_color='E6A5AC',
                                            end_color='E6A5AC',
                                            fill_type='solid')

                # 获取列索引
                header = [cell.value for cell in worksheet[1]]
                col_indices = {col: idx + 1 for idx, col in enumerate(header)}

                has_consistency = '结果是否一致' in col_indices
                has_diff = '差异性' in col_indices

                # 遍历数据行应用样式
                for row_idx in range(2, worksheet.max_row + 1):
                    fill_color = None

                    if has_consistency:
                        consistency = worksheet.cell(
                            row=row_idx, column=col_indices['结果是否一致']).value

                        if consistency == '一致':
                            fill_color = green_fill
                        else:
                            # 如果不一致，根据差异程度细分颜色 (或者简单均使用红色)
                            if has_diff:
                                difference = worksheet.cell(
                                    row=row_idx,
                                    column=col_indices['差异性']).value
                                if difference == '差两级以上':
                                    fill_color = dark_red_fill
                                else:
                                    fill_color = red_fill
                            else:
                                fill_color = red_fill

                    # 如果确定了颜色，应用到整行
                    if fill_color:
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx,
                                           column=col_idx).fill = fill_color

                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            val_len = len(str(cell.value)) if cell.value else 0
                            if val_len > max_length:
                                max_length = val_len
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    # 限制最大宽度，防止过宽
                    worksheet.column_dimensions[column_letter].width = min(
                        adjusted_width, 50)

            QMessageBox.information(self, "保存成功", f"结果已保存至: {output_file}")

        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存过程中发生错误: {str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
