import sys
import pandas as pd
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication, QWidget, QTreeWidgetItem, QFileDialog,
    QTableWidgetItem, QMessageBox, QTreeWidget, QVBoxLayout,
    QPushButton, QHBoxLayout, QTableWidget, QComboBox, QLabel
)
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


class StudentAidPlugin(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("家庭经济困难学生精准认定插件")
        self.setMinimumSize(900, 600)
        self.setStyleSheet("""QWidget { background-color: #f0f2f5; font-family: 'Microsoft YaHei'; font-size: 13px; }
        QPushButton { background-color: #3f72af; color: white; border-radius: 5px; padding: 8px 15px; font-weight: 600; min-width: 150px; }
        QPushButton:hover { background-color:  #2b5d8c; }
        QPushButton:pressed { background-color: #3071e7; }
        QTreeWidget, QTableWidget { background-color: white; border: 1px solid #dcdfe6; border-radius: 4px; }
        QTableWidget { selection-background-color: #c6e2ff; gridline-color: #ebeef5; }
        QLabel { font-size: 12px; color: #444; }
        """)

        self.indicator_df = None
        self.selected_indicators = []

        layout = QHBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(15, 15, 15, 15)

        # 左侧：指标体系区域
        leftWidget = QWidget()
        leftLayout = QVBoxLayout(leftWidget)
        leftLayout.setSpacing(12)
        leftLayout.setContentsMargins(10, 10, 10, 10)

        importLayout = QHBoxLayout()
        self.loadIndicatorButton = QPushButton("导入指标体系文件")
        self.filePathLabel = QLabel("未选择文件")
        importLayout.addWidget(self.loadIndicatorButton)
        importLayout.addWidget(self.filePathLabel)

        self.tipLabel = QLabel("提示：选择指标后可生成模板，右侧录入学生信息可导出数据。")

        exportTemplateLayout = QHBoxLayout()
        self.generateButton = QPushButton("导出学生数据模板")
        self.templatePathLabel = QLabel("未生成模板")
        exportTemplateLayout.addWidget(self.generateButton)
        exportTemplateLayout.addWidget(self.templatePathLabel)

        self.indicatorTreeWidget = QTreeWidget()
        self.indicatorTreeWidget.setMinimumHeight(350)
        self.indicatorTreeWidget.setHeaderLabels(["指标选择"])

        leftLayout.addLayout(importLayout)
        leftLayout.addWidget(self.tipLabel)  # 提示标签放在按钮上面

        leftLayout.addWidget(self.indicatorTreeWidget)
        leftLayout.addLayout(exportTemplateLayout)  # 按钮放下面


        # 右侧：学生数据操作区
        rightWidget = QWidget()
        rightLayout = QVBoxLayout(rightWidget)
        rightLayout.setSpacing(8)
        rightLayout.setContentsMargins(0, 0, 0, 0)

        # 修改后的按钮布局：仅包含新增和删除按钮
        btnLayout = QHBoxLayout()
        self.addStudentButton = QPushButton("新增学生")
        self.deleteStudentButton = QPushButton("删除选中学生")
        btnLayout.addWidget(self.addStudentButton)
        btnLayout.addWidget(self.deleteStudentButton)

        self.indicatorTableWidget = QTableWidget()
        self.indicatorTableWidget.setColumnCount(0)
        self.indicatorTableWidget.setRowCount(0)
        self.indicatorTableWidget.setMinimumWidth(520)

        # 新增：导出按钮及路径显示独立在表格下方
        exportLayout = QHBoxLayout()
        self.exportDataButton = QPushButton("导出学生数据")
        self.exportPathLabel = QLabel("未导出文件")
        exportLayout.addWidget(self.exportDataButton)
        exportLayout.addWidget(self.exportPathLabel)

        rightLayout.addLayout(btnLayout)
        rightLayout.addWidget(self.indicatorTableWidget)
        rightLayout.addLayout(exportLayout)

        layout.addWidget(leftWidget,1)
        layout.addWidget(rightWidget,2)

        self.loadIndicatorButton.clicked.connect(self.load_indicator_file)
        self.generateButton.clicked.connect(self.generate_template)
        self.indicatorTreeWidget.itemChanged.connect(self.handle_item_changed)
        self.addStudentButton.clicked.connect(self.add_student_row)
        self.deleteStudentButton.clicked.connect(self.delete_selected_students)
        self.exportDataButton.clicked.connect(self.export_student_data)

    def load_indicator_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择指标体系文件", "", "CSV文件 (*.csv);;Excel文件 (*.xlsx)")
        if not file_path:
            return
        try:
            if file_path.endswith('.csv'):
                self.indicator_df = pd.read_csv(file_path)
            else:
                self.indicator_df = pd.read_excel(file_path)
            self.filePathLabel.setText(file_path)
            self.build_indicator_tree()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载指标文件失败：{e}")

    def build_indicator_tree(self):
        self.indicatorTreeWidget.blockSignals(True)
        self.indicatorTreeWidget.clear()
        self.selected_indicators.clear()
        self.indicator_options = {}

        # 记录二级指标顺序，方便后续排序
        self.indicator_order = {}
        # 用 drop_duplicates 保持指标_2的原始出现顺序，并赋序号
        unique_ind2 = self.indicator_df['indicator_2'].drop_duplicates()
        for idx, ind2 in enumerate(unique_ind2):
            self.indicator_order[ind2] = idx

        self.indicator_df['__order__'] = range(len(self.indicator_df))
        grouped = self.indicator_df.sort_values('__order__').groupby(['indicator_1', 'indicator_2'])

        root_items = {}
        for (ind1, ind2), group in grouped:
            if ind1 not in root_items:
                item1 = QTreeWidgetItem(self.indicatorTreeWidget)
                item1.setText(0, ind1)
                item1.setFlags(item1.flags() | Qt.ItemIsUserCheckable)
                item1.setCheckState(0, Qt.Unchecked)
                root_items[ind1] = item1
            item1 = root_items[ind1]

            item2 = QTreeWidgetItem(item1)
            item2.setText(0, ind2)
            item2.setFlags(item2.flags() | Qt.ItemIsUserCheckable)
            item2.setCheckState(0, Qt.Unchecked)

            third_list = list(group['indicator_3'].dropna().unique())
            self.indicator_options[ind2] = third_list

            for _, row in group.iterrows():
                item3 = QTreeWidgetItem(item2)
                item3.setText(0, row['indicator_3'])
                item3.setFlags(item3.flags() & ~Qt.ItemIsUserCheckable)

        self.indicatorTreeWidget.expandAll()
        self.indicatorTreeWidget.blockSignals(False)

    def export_student_data(self):
        if self.indicatorTableWidget.rowCount() == 0:
            QMessageBox.information(self, "提示", "暂无学生数据可导出。")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "导出学生数据", "", "Excel 文件 (*.xlsx)")
        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            headers = [self.indicatorTableWidget.horizontalHeaderItem(i).text() for i in range(self.indicatorTableWidget.columnCount())]
            ws.append(headers)

            for row in range(self.indicatorTableWidget.rowCount()):
                row_data = []
                for col in range(self.indicatorTableWidget.columnCount()):
                    if col == 0:
                        item = self.indicatorTableWidget.item(row, col)
                        row_data.append(item.text() if item else "")
                    else:
                        widget = self.indicatorTableWidget.cellWidget(row, col)
                        row_data.append(widget.currentText() if widget else "")
                ws.append(row_data)

            wb.save(save_path)
            # self.exportPathLabel.setText(save_path)
            QMessageBox.information(self, "导出成功", f"学生数据已导出到：\n{save_path}")
            self.exportPathLabel.setText(save_path)


            reply = QMessageBox.question(
                self,
                "清空数据",
                "是否清空已显示的学生数据表，重新选择指标？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.clear_student_data_table()

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出时出错:\n{str(e)}")

    # 其他方法不变，如 handle_item_changed, refresh_indicator_table, add_student_row, delete_selected_students, generate_template, clear_student_data_table 等



    # 其他方法不变，如 handle_item_changed, refresh_indicator_table, add_student_row, delete_selected_students, generate_template, clear_student_data_table 等

    def handle_item_changed(self, item, column):
        if column != 0:
            return

        if self.indicatorTableWidget.rowCount() > 0:
            QMessageBox.warning(self, "操作禁止", "已有学生数据，请先清空或导出数据后再修改指标体系。")
            self.indicatorTreeWidget.blockSignals(True)
            item.setCheckState(0, Qt.Unchecked if item.checkState(0) == Qt.Checked else Qt.Checked)
            self.indicatorTreeWidget.blockSignals(False)
            return

        self.indicatorTreeWidget.blockSignals(True)
        try:
            state = item.checkState(0)
            parent = item.parent()

            def add_indicator(ind2):
                if ind2 not in self.selected_indicators:
                    self.selected_indicators.append(ind2)
                    # 按顺序排序
                    self.selected_indicators.sort(key=lambda x: self.indicator_order.get(x, 9999))

            if parent is None:
                for i in range(item.childCount()):
                    child = item.child(i)
                    child.setCheckState(0, state)
                    ind2 = child.text(0)
                    if state == Qt.Checked:
                        add_indicator(ind2)
                    else:
                        if ind2 in self.selected_indicators:
                            self.selected_indicators.remove(ind2)
            else:
                ind2 = item.text(0)
                if state == Qt.Checked:
                    add_indicator(ind2)
                else:
                    if ind2 in self.selected_indicators:
                        self.selected_indicators.remove(ind2)

                checked = sum(parent.child(i).checkState(0) == Qt.Checked for i in range(parent.childCount()))
                if checked == parent.childCount():
                    parent.setCheckState(0, Qt.Checked)
                elif checked == 0:
                    parent.setCheckState(0, Qt.Unchecked)
                else:
                    parent.setCheckState(0, Qt.PartiallyChecked)

            self.refresh_indicator_table()
        finally:
            self.indicatorTreeWidget.blockSignals(False)

    def refresh_indicator_table(self):
        table = self.indicatorTableWidget
        table.clear()
        table.setRowCount(0)
        # 直接使用列表，保持添加顺序
        headers = ["学生姓名"] + self.selected_indicators
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        table.resizeColumnsToContents()
        for col in range(len(headers)):
            if table.columnWidth(col) < 100:
                table.setColumnWidth(col, 100)

    def add_student_row(self):
        row = self.indicatorTableWidget.rowCount()
        self.indicatorTableWidget.insertRow(row)
        self.indicatorTableWidget.setItem(row, 0, QTableWidgetItem(""))  # 姓名列

        for col in range(1, self.indicatorTableWidget.columnCount()):
            indicator_name = self.indicatorTableWidget.horizontalHeaderItem(col).text()
            combo = QComboBox()
            options = self.indicator_options.get(indicator_name, ["未定义"])
            combo.addItems(options)
            self.indicatorTableWidget.setCellWidget(row, col, combo)

    def delete_selected_students(self):
        selected_rows = set(i.row() for i in self.indicatorTableWidget.selectedItems())
        for row in sorted(selected_rows, reverse=True):
            self.indicatorTableWidget.removeRow(row)

    from PyQt5.QtWidgets import QMessageBox



    def generate_template(self):
        if not self.selected_indicators:
            QMessageBox.warning(self, "提示", "请至少选择一个二级指标")
            return

        save_dir = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if not save_dir:
            QMessageBox.warning(self, "提示", "请选择一个文件夹")
            return

        QMessageBox.information(self, "提示", "请确保未打开之前生成的“学生输入数据模版.xlsx”文件，否则可能导致保存失败。")
        try:
            sorted_indicators = sorted(self.selected_indicators, key=lambda x:
                list(self.indicator_df[self.indicator_df['indicator_2'] == x].index)[0])
            selected_df = self.indicator_df[self.indicator_df['indicator_2'].isin(sorted_indicators)].copy()
            selected_df.reset_index(drop=True, inplace=True)
            selected_df.insert(0, "序号", selected_df.index + 1)
            selected_df.to_csv(f"{save_dir}/选择后的指标体系.csv", index=False, encoding="utf-8-sig")

            indicator_3_map = {}
            for ind2 in sorted_indicators:
                options = selected_df[selected_df['indicator_2'] == ind2]['indicator_3'].unique().tolist()
                indicator_3_map[ind2] = options

            wb = Workbook()
            ws = wb.active
            ws.title = "学生数据输入模板"
            all_ind2 = self.indicator_df['indicator_2']
            seen = set()
            ordered = [i for i in all_ind2 if not (i in seen or seen.add(i))]
            ordered_selected = [i for i in ordered if i in self.selected_indicators]
            headers = ["学生姓名"] + ordered_selected
            ws.append(headers)

            for col_idx, ind2 in enumerate(ordered_selected, start=2):
                options = ",".join(indicator_3_map[ind2])
                dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
                ws.add_data_validation(dv)
                col_letter = get_column_letter(col_idx)
                dv.add(f"{col_letter}2:{col_letter}1048576")

            wb.save(f"{save_dir}/学生输入数据模版.xlsx")
            self.templatePathLabel.setText(f"{save_dir}/学生输入数据模版.xlsx")
            QMessageBox.information(self, "成功", "模板已成功生成，含三级指标下拉菜单（xlsx格式）")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成模板失败：{e}")
    def clear_student_data_table(self):
        # 保留表头，清空所有行数据
        self.indicatorTableWidget.setRowCount(0)




if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = StudentAidPlugin()
    win.show()
    sys.exit(app.exec_())